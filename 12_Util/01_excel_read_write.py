print('-start excel pre-processing-')

#참고: https://myjamong.tistory.com/51
from openpyxl import load_workbook

# data_only=Ture로 해줘야 수식이 아닌 값으로 받아온다.
data_path = ""
file_name = ""

load_wb = load_workbook(data_path + file_name, data_only=True)
# 시트 이름으로 불러오기
load_ws = load_wb['']

print('\n-----모든 행과 열 출력-----')
cnt = 0
for row_idx, row in enumerate(load_ws.rows):
    row_value = []
    for col_idx, cell in enumerate(row):
        #1)첫번째 열이 숫자가 아니면 제외
        print(type(cell.value))
        if cell.value != None:
            print('row_idx:', row_idx, '/col_idx:', col_idx, 'val:', cell.value)
        if col_idx == 0 :
            if cell.value == None : continue
            if len(cell.value) == 0 : continue
            #cel_val = int(cell.value)
            #print('row_idx:', row_idx, '/col_idx:', col_idx, 'val:', cell.value, ',', cell.value.isdigit(), ',', cel_val)
            if cell.value.isdigit() == False: continue
        if cell.value != None :
           #print('row_idx:', row_idx, '/col_idx:', col_idx, 'val:', cell.value, ',', cell.value.isdigit())
            print(cell.value)
    if cnt > 100: break
    cnt += 1

print('-end excel pre-processing-')

# from openpyxl import Workbook
# write_wb = Workbook()
# # 이름이 있는 시트를 생성
# # write_ws = write_wb.create_sheet('생성시트')
# # Sheet1에다 입력
# write_ws = write_wb.active
#
# # 행 단위로 추가
# write_ws.append([1, 2, 3])
# save_file_name = "2017품목분류표(HSK)_DATA추출.xlsx"
# write_wb.save(data_path + save_file_name)
#
# print('-end excel save')